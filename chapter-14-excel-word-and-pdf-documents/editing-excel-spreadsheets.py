import openpyxl
import os

cwd = f'{os.getcwd()}/chapter-14-excel-word-and-pdf-documents'

wb = openpyxl.Workbook()
print(wb.sheetnames)

sheet = wb['Sheet']
print(sheet['A1'].value)

sheet['A1'].value = 42
sheet['A2'].value = 'Hello'

wb.save(f'{cwd}/editing1.xlsx')
sheet2 = wb.create_sheet()
sheet2.title = 'My New Sheet Name'
wb.save(f'{cwd}/editing2.xlsx')
print(wb.sheetnames)

wb.create_sheet(index=0, title='My Other Sheet')
wb.save(f'{cwd}/editing3.xlsx')
print(wb.sheetnames)
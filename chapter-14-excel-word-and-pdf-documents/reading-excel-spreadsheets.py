import openpyxl
import os

cwd = f'{os.getcwd()}/chapter-14-excel-word-and-pdf-documents'

workbook = openpyxl.load_workbook(f'{cwd}/example.xlsx')
print(type(workbook))

print(workbook.sheetnames)
sheet = workbook['Sheet1']
print(type(sheet))

cell = sheet['A1']
print(sheet.cell(row=1, column=1).value)
print(cell.value)

cell = sheet['B1']
print(sheet.cell(row=1, column=2).value)
print(cell.value)

cell = sheet['C1']
print(sheet.cell(row=1, column=3).value)
print(cell.value)

for i in range(1, 8):
	print(i ,sheet.cell(row=1, column=2).value)